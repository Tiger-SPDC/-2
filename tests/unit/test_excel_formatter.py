"""ExcelFormatter 单元测试：8 个 Sheet、标题行、字节输出（全离线）。"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from industry_intelligence.reporting.builder import ReportDataBundle
from industry_intelligence.reporting.formatters.excel import ExcelFormatter

_EXPECTED_SHEETS = [
    "Run_Summary", "Events", "Companies", "Metrics", "Documents",
    "Claims", "Evidence", "Data_Quality",
]

_CLAIM = {
    "claim_id": "c1",
    "claim_text": "特来电市占率上升",
    "claim_type": "fact",
    "confidence": 0.9,
    "entity_id": "特来电",
    "analysis_type": "market",
    "evidence": [{"document_id": "d1", "evidence_role": "primary_source"}],
}


def _bundle(**overrides) -> ReportDataBundle:
    kwargs: dict[str, object] = dict(
        run_id="r1",
        topic_id="t1",
        task_id="tk1",
        status="success",
        period_start="2026-01-01",
        period_end="2026-01-08",
        events=[{"event_id": "e1", "event_type_id": "new_product",
                 "event_date": "2026-01-05", "title": "发布新品", "confidence": 0.9}],
        companies=[{"name": "特来电", "aliases": ["特来电新能源"]}],
        observations=[{"metric_id": "station_count", "entity_id": "特来电",
                       "value": 100.0, "unit": "座", "period_end": "2026-01-08",
                       "confidence": 0.9}],
        documents=[{"document_id": "d1", "title": "文档一", "source_id": "rss:demo",
                    "source_grade": "C", "fetched_at": "2026-01-05"}],
        claims=[_CLAIM],
        review_results=[{"claim_id": "c1", "verdict": "pass", "issues": []}],
        quality={
            "document_count": 1.0, "event_count": 1.0, "observation_count": 1.0,
            "claim_count": 1.0, "evidence_coverage": 1.0,
            "review_count": 1.0, "review_reject_count": 0.0,
        },
    )
    kwargs.update(overrides)
    return ReportDataBundle(**kwargs)


def _load(bundle: ReportDataBundle):
    data = ExcelFormatter().render(bundle)
    assert isinstance(data, bytes)
    assert data[:2] == b"PK"  # xlsx zip magic
    return load_workbook(BytesIO(data))


def test_render_produces_all_eight_sheets() -> None:
    wb = _load(_bundle())
    assert wb.sheetnames == _EXPECTED_SHEETS


def test_run_summary_has_kv_rows() -> None:
    wb = _load(_bundle())
    ws = wb["Run_Summary"]
    pairs = {(ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value)
             for r in range(1, ws.max_row + 1)}
    assert ("run_id", "r1") in pairs
    assert ("status", "success") in pairs


def test_claims_sheet_links_review_verdict() -> None:
    wb = _load(_bundle())
    ws = wb["Claims"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "review_verdict" in headers
    # 标题行加粗 + 冻结首行
    assert ws.freeze_panes == "A2"
    assert ws.cell(row=1, column=1).font.bold


def test_evidence_sheet_expands_claim_evidence() -> None:
    wb = _load(_bundle())
    ws = wb["Evidence"]
    # 第 2 行为 c1 的文档证据
    assert ws.cell(row=2, column=1).value == "c1"
    assert ws.cell(row=2, column=2).value == "d1"
    assert ws.cell(row=2, column=4).value == "primary_source"


def test_empty_bundle_still_renders() -> None:
    bundle = _bundle(
        events=[], companies=[], observations=[], documents=[], claims=[],
        review_results=[], quality={"document_count": 0.0},
    )
    wb = _load(bundle)
    assert wb["Events"].max_row == 1  # 只有标题行
    assert wb["Data_Quality"].max_row >= 2
