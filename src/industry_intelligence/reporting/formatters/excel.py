"""Excel 周报（Phase 4，§20.3）：8 个 Sheet 的结构化数据导出。

使用 openpyxl 直接写入（不依赖 pandas）。所有内容来自 ReportDataBundle，
行业数据（公司名/指标名等）全部来自 bundle，不在代码中硬编码。
"""

from __future__ import annotations

from collections.abc import Sequence
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from industry_intelligence.reporting.builder import ReportDataBundle

_HEADER_FONT = Font(bold=True)
_SHEET_MAX_WIDTH = 60


def _join_list(value: object) -> str:
    """把 bundle 里可能是 list 的值安全 join 成字符串。"""
    if not isinstance(value, list):
        return ""
    return "；".join(str(x) for x in value)


class ExcelFormatter:
    """把 ReportDataBundle 渲染为 8-Sheet xlsx 字节。"""

    def render(self, bundle: ReportDataBundle) -> bytes:
        wb = Workbook()
        self._run_summary(wb, bundle)
        self._events(wb, bundle)
        self._companies(wb, bundle)
        self._metrics(wb, bundle)
        self._documents(wb, bundle)
        self._claims(wb, bundle)
        self._evidence(wb, bundle)
        self._data_quality(wb, bundle)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------- Sheet

    def _run_summary(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.active
        assert isinstance(ws, Worksheet)  # Workbook() 默认含一个 Worksheet
        ws.title = "Run_Summary"
        rows = [
            ("run_id", bundle.run_id),
            ("topic_id", bundle.topic_id),
            ("task_id", bundle.task_id),
            ("status", bundle.status),
            ("period_start", bundle.period_start),
            ("period_end", bundle.period_end),
            ("document_count", bundle.quality.get("document_count", 0)),
            ("event_count", bundle.quality.get("event_count", 0)),
            ("observation_count", bundle.quality.get("observation_count", 0)),
            ("claim_count", bundle.quality.get("claim_count", 0)),
            ("evidence_coverage", bundle.quality.get("evidence_coverage", 0)),
            ("review_count", bundle.quality.get("review_count", 0)),
            ("review_reject_count", bundle.quality.get("review_reject_count", 0)),
        ]
        self._write_sheet(ws, rows, headers=("字段", "值"))

    def _events(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Events")
        headers = ("event_id", "event_type_id", "event_date", "title", "confidence")
        rows = [
            (
                e.get("event_id", ""),
                e.get("event_type_id", ""),
                e.get("event_date", ""),
                e.get("title", ""),
                e.get("confidence", 0.0),
            )
            for e in bundle.events
        ]
        self._write_sheet(ws, rows, headers=headers)

    def _companies(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Companies")
        headers = ("公司", "别名")
        rows = [
            (c.get("name", ""), _join_list(c.get("aliases")))
            for c in bundle.companies
        ]
        self._write_sheet(ws, rows, headers=headers)

    def _metrics(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Metrics")
        headers = ("metric_id", "entity_id", "value", "unit", "period_end", "confidence")
        rows = [
            (
                o.get("metric_id", ""),
                o.get("entity_id", ""),
                o.get("value", 0.0),
                o.get("unit", ""),
                o.get("period_end", ""),
                o.get("confidence", 0.0),
            )
            for o in bundle.observations
        ]
        self._write_sheet(ws, rows, headers=headers)

    def _documents(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Documents")
        headers = ("document_id", "title", "source_id", "source_grade", "published_at")
        rows = [
            (
                d.get("document_id", ""),
                d.get("title", ""),
                d.get("source_id", ""),
                d.get("source_grade", ""),
                d.get("published_at", d.get("fetched_at", "")),
            )
            for d in bundle.documents
        ]
        self._write_sheet(ws, rows, headers=headers)

    def _claims(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Claims")
        headers = (
            "claim_id", "claim_text", "claim_type", "confidence",
            "entity_id", "analysis_type", "review_verdict", "issues",
        )
        rows: list[tuple[object, ...]] = []
        for c in bundle.claims:
            review = self._review_for(bundle, str(c.get("claim_id", "")))
            rows.append(
                (
                    c.get("claim_id", ""),
                    c.get("claim_text", ""),
                    c.get("claim_type", ""),
                    c.get("confidence", 0.0),
                    c.get("entity_id", ""),
                    c.get("analysis_type", ""),
                    review.get("verdict", "") if review else "",
                    _join_list(review.get("issues")) if review else "",
                )
            )
        self._write_sheet(ws, rows, headers=headers)

    def _evidence(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Evidence")
        headers = ("claim_id", "document_id", "observation_id", "evidence_role")
        rows: list[tuple[object, ...]] = []
        for c in bundle.claims:
            evidence = c.get("evidence")
            if not isinstance(evidence, list):
                continue
            for ev in evidence:
                rows.append(
                    (
                        c.get("claim_id", ""),
                        ev.get("document_id", "") or "",
                        ev.get("observation_id", "") or "",
                        ev.get("evidence_role", ""),
                    )
                )
        self._write_sheet(ws, rows, headers=headers)

    def _data_quality(self, wb: Workbook, bundle: ReportDataBundle) -> None:
        ws = wb.create_sheet("Data_Quality")
        rows = [(k, v) for k, v in bundle.quality.items()]
        self._write_sheet(ws, rows, headers=("指标", "值"))

    # ------------------------------------------------------------- 助手

    def _write_sheet(
        self,
        ws: Worksheet,
        rows: Sequence[Sequence[Any]],
        *,
        headers: tuple[str, ...],
    ) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append(tuple(row))
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(text))
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                min(_SHEET_MAX_WIDTH, max_len)
            )

    def _review_for(
        self, bundle: ReportDataBundle, claim_id: str
    ) -> dict[str, object] | None:
        for r in bundle.review_results:
            if str(r.get("claim_id", "")) == claim_id:
                return r
        return None
